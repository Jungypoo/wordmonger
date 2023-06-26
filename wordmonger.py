import requests, sys, datamuse, traceback
from bs4 import BeautifulSoup
from openpyxl import Workbook, descriptors

args = sys.argv[1:]
dm = datamuse.Datamuse()

# Creating a sheet in the workbook for each keyword
wb = Workbook()
for i in range(len(args)):
    wb.create_sheet(args[i], i)

# Getting some synonyms
def get_synonyms(word):
    listy = []
    synonyms = dm.words(rel_syn=word)
    means_like = dm.words(ml=word)
    
    for i in synonyms:
        listy.append(i['word'])

    for i in means_like:
        listy.append(i['word'])
    
    listy = list(dict.fromkeys(listy))

    for i in reversed(range(len(listy))):
        if word in listy[i]:
            del listy[i]

    return listy

# If noun, serve popular adjectives. If adj, serve popular nouns
def get_modifiers(word):
    lista = []
    listb = []
    mod_jja = dm.words(rel_jja=word)
    mod_jjb = dm.words(rel_jjb=word)

    for i in mod_jja:
        lista.append(i['word'])

    for i in mod_jjb:
        listb.append(i['word'])

    return lista, listb

# Trigger words
def get_triggers(word):
    listy = []
    triggers = dm.words(rel_trg=word)

    for i in triggers:
        listy.append(i['word'])

    return listy

# Holonyms and meronyms
def get_holynyms(word):
    listy = []
    holonyms = dm.words(rel_com=word)
    for i in holonyms:
        listy.append(i['word'])
    return listy

def get_meronyms(word):
    listy = []
    meronyms = dm.words(rel_par=word)
    for i in meronyms:
        listy.append(i['word'])
    return listy

# Followers and predecessors
def get_predecessors(word):
    listy = []
    predecessors = dm.words(rel_bgb=word)
    for i in predecessors:
        listy.append(i['word'])
    return listy

def get_followers(word):
    listy = []
    followers = dm.words(rel_bga=word)
    for i in followers:
        listy.append(i['word'])
    return listy

# Rhymes with
def get_rhymes(word):
    listy = []
    rhymes = dm.words(rel_rhy=word)
    for i in rhymes:
        listy.append(i['word'])
    return listy

# Consonant match and homophones
def get_soundslike(word):
    listy = []
    consonant_match = dm.words(rel_cns=word)
    homophones = dm.words(rel_hom=word)
    for i in homophones:
        consonant_match.append(i)

    for i in consonant_match:
        listy.append(i['word'])

    listy = list(dict.fromkeys(listy))

    return listy

# Idioms
def get_idioms(word):
    listy = []
    idioms = dm.words(sp= '*' + word + '*')

    # Discount the too-short results
    for i in reversed(range(len(idioms))):
        if idioms[i]['word'].find(' ') == -1:
            del idioms[i]

    for i in idioms:
        listy.append(i['word'])

    return listy

# Loop through the arguments, call all functions, and write
for i in args:

    ws = wb[i]

    try:
        # Writing synonyms
        synonyms = get_synonyms(i)
        ws['A1'] = 'Synonyms'
        for x in range(len(synonyms)):
            ws.cell(row=x+2, column=1).value = synonyms[x]
    except Exception:
        print('\nProblem writing synonyms:')
        print(traceback.format_exc())

    try:
        # Modifiers for nouns and adjectives (both ways)
        mod_jja, mod_jjb = get_modifiers(i)
        ws['B1'] = 'noun -> adj'
        ws['C1'] = 'adj -> noun'
        for x in range(len(mod_jja)):
            ws.cell(row=x+2, column=2).value = mod_jja[x]
        for x in range(len(mod_jjb)):
            ws.cell(row=x+2, column=3).value = mod_jjb[x]
    except Exception:
        print('\nProblem writing modifiers:')
        print(traceback.format_exc())

    try:
        # Triggers, words associated with there query word in the same piece of text
        triggers = get_triggers(i)
        ws['D1'] = 'Triggers'
        for x in range(len(triggers)):
            ws.cell(row=x+2, column=4).value = triggers[x]
    except Exception:
        print('\nProblem writing triggers:')
        print(traceback.format_exc())
    
    try:
        # Holonyms and Meronyms
        holonyms = get_holynyms(i)
        ws['E1'] = 'Holonyms'
        for x in range(len(holonyms)):
            ws.cell(row=x+2, column=5).value = holonyms[x]

        meronyms = get_meronyms(i)
        ws['F1'] = 'Meronyms'
        for x in range(len(meronyms)):
            ws.cell(row=x+2, column=6).value = meronyms[x]

    except Exception:
        print('\nProblem writing holonyms and meronyms:')
        print(traceback.format_exc())

    try:
        # Predecessors and Followers
        predecessors = get_predecessors(i)
        ws['G1'] = 'Predecessors'
        for x in range(len(predecessors)):
            ws.cell(row=x+2, column=7).value = predecessors[x]

        followers = get_followers(i)
        ws['H1'] = 'Followers'
        for x in range(len(followers)):
            ws.cell(row=x+2, column=8).value = followers[x]

    except Exception:
        print('\nProblem writing predecessors and followers:')
        print(traceback.format_exc())

    try:
        # Rhymes
        rhymes = get_rhymes(i)
        ws['I1'] = 'Rhymes'
        for x in range(len(rhymes)):
            ws.cell(row=x+2, column=9).value = rhymes[x]
    except Exception:
        print('\nProblem writing rhymes:')
        print(traceback.format_exc())

    try:
        # Sounds like...
        soundslike = get_soundslike(i)
        ws['J1'] = 'Sounds like...'
        for x in range(len(soundslike)):
            ws.cell(row=x+2, column=10).value = soundslike[x]
    except Exception:
        print('\nProblem writing Soundslike:')
        print(traceback.format_exc())

    try:
        # Idioms
        idioms = get_idioms(i)
        ws['K1'] = 'Idioms'
        for x in range(len(idioms)):
            ws.cell(row=x+2, column=11).value = idioms[x]
    except Exception:
        print('\nProblem writing idioms:')
        print(traceback.format_exc())

# Make the columns a bit wider
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 20
ws.column_dimensions['H'].width = 20
ws.column_dimensions['I'].width = 20
ws.column_dimensions['J'].width = 20
ws.column_dimensions['K'].width = 20

# Write to spreadsheet
save_string = ''
for i in args:
    save_string += ' ' + i

wb.save(save_string + '.xlsx')
